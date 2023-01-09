from datetime import datetime, date
import json
from openpyxl import load_workbook
import os
import requests
import time
import shutil
import sys
#import unicodedata

import myGlobals as mg
import httpHeaders as hh
import config

from common.utils import myprint, color, dumpToFile, dumpJsonToFile, dumpListToFile, dumpListOfListToFile, bubbleSort, isFileOlderThanXMinutes

class color:
    PURPLE    = '\033[95m'
    CYAN      = '\033[96m'
    DARKCYAN  = '\033[36m'
    BLUE      = '\033[94m'
    GREEN     = '\033[92m'
    YELLOW    = '\033[93m'
    RED       = '\033[91m'
    BOLD      = '\033[1m'
    GREYED    = '\033[2m'
    ITALIC    = '\033[3m'
    UNDERLINE = '\033[4m'
    STRIKETHRu = '\033[9m'
    END       = '\033[0m'


# Dictionary containing the HTTP requests to send to the server

#https://www.mcb.mu/fr/handler/GetForexData?fxc=EUR&_=1671555466629

MCB_URL = 'https://www.mcb.mu/fr/personal/download-daily-rates'

MCB_DAILY_RATES_HTTP_REQUESTS = {
    "dailyRatesForEuro" : {
        "name" : "dailyRatesForEuro",
        "info" : "Connect to MCB.mu and get daily rates for Euro currency",
        "rqst" : {
            "type" : 'POST',
            "url"  : MCB_URL,
            "headers" : {},
            "payload_type" : 'MULTIPART_FORM_DATA',
            "payload_data" : {
                'forexCurrency': (None, 'EUR'),
                'forexDateStart': (None, 'PLACEHOLDER-FOR-START-DATE'),  # 19/12/2022
                'forexDateEnd': (None, 'PLACEHOLDER-FOR-END-DATE'),    # 19/12/2022
                '__RequestVerificationToken': (None, 'RLOHjJ_oH627NJuu1v6VqTGvrSNNV-q5X0PH5fhUexGCyl0clAurWm0byFPSMLLCsFGixmBTBqB2KJ6i1ALum6XeEjRZSesIDq_yyasBmTs1'),
                'tprt': (None, '2FAEB4E8743372FCE42814E9A6DF587B29B7907D5B488599FC9BFB4BC71132104CD4D509BB480A16CC14F25F9AA42BE2FD1E1AE322F7802814335494FA64760E5C1CB2386E8F1A8D01D4FC22BC5A9AD854FF4E669C9D036F98CFD7435BBC4D07D34957F072FC588BB83C0EF88A11CBD871FEA398A66B71CB6D24AB035867F7BBE73A69F862F40A045E04511C6A001A673DFD31E991FCA7536992460D2ABDC416C3E35D6065C0A357621783657131DC69C7974ABFCB37D1051AA4A1A81B53E490'),
            }
        },
        "resp" : {
            "code" : 200,
            "dumpResponse" : 'foo.xlsx',
            # If Content-Disposition is present in response header, use it to get output filename
            "useContentDisposition" : True,
            "updateCookies" : False,
        },
        "returnText" : False,
    },
}
    
cacheUpdated = False

class MCBDailyRates:
    def __init__(self, session):
        self._session  = session
        # Dict to save cookies from server
        self._cookies = dict()

    def getDailyRates(self, forexDate):  # dd/mm/yyyy
        # Update the 'dailyRatesForEuro' request with the requested date
        httpRqst = MCB_DAILY_RATES_HTTP_REQUESTS['dailyRatesForEuro']
        myprint(2, httpRqst)
        httpRqst['rqst']['payload_data']['forexDateStart'] = (None, forexDate)
        httpRqst['rqst']['payload_data']['forexDateEnd']   = (None, forexDate)

        dt = datetime.strptime(forexDate,'%d/%m/%Y').strftime('%Y%m%d')
        httpRqst['resp']['dumpResponse'] = "%s.xlsx" % (dt)
        
        myprint(2, 'Using HTTP Request:', httpRqst)
        
        # Execute request to get the daily rates information
        respText = self._executeRequest(httpRqst)
        if 'ErRoR' in respText:
            myprint(1, 'Error retrieving information from server')
            return -1

        # Parse returned information. Create/Update local cache file
        parseDailyRates(httpRqst['resp']['dumpResponse'], dt)
        if not config.KEEPRESPONSEFILE:
            myprint(1, 'Removing: %s' % (httpRqst['resp']['dumpResponse']))
            os.remove(httpRqst['resp']['dumpResponse'])
            
        return 0
        
    # Build a string containing all cookies passed as parameter in a list 
    def _buildCookieString(self, cookieList):
        cookieAsString = ''
        for c in cookieList:
            # Check if cookie exists in our dict
            if c in self._cookies:
                cookieAsString += '%s=%s; ' % (c, self._cookies[c])
            else:
                myprint(1,'Warning: Cookie %s not found.' % (c))
        return(cookieAsString)

    # Update our cookie dict
    def _updateCookies(self, cookies):
        for cookie in self._session.cookies:
            if cookie.value == 'undefined' or cookie.value == '':
                myprint(2,'Skipping cookie with undefined value %s' % (cookie.name))
                continue
            if cookie.name in self._cookies and self._cookies[cookie.name] != cookie.value:
                myprint(1,'Updating cookie:', cookie.name)
                self._cookies[cookie.name] = cookie.value
            elif not cookie.name in self._cookies:
                myprint(1,'Adding cookie:', cookie.name)
                self._cookies[cookie.name] = cookie.value
            else:
                myprint(2,'Cookie not modified:', cookie.name)                

    def _executeRequest(self, rqst):
        dt_now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        myprint(1, '%s: Executing request "%s": %s' % (dt_now, rqst["name"], rqst["info"]))
        myprint(2, json.dumps(rqst, indent=4))

        hdrs = hh.HttpHeaders()

        for k,v in rqst["rqst"]["headers"].items():
            if k == "Cookie":
                if 'str' in str(type(v)):	# Cookie is a string
                    cookieAsString = v
                else:				# Cookie is a list of cookies
                    assert('list' in str(type(v)))
                    cookieAsString = self._buildCookieString(v)

                # Add extra Cookie if requested
                if "extraCookie" in rqst["rqst"]:
                    cookieAsString += rqst["rqst"]["extraCookie"]
                hdrs.setHeader('Cookie', cookieAsString)
            else:
                hdrs.setHeader(k, v)

        rqstType = rqst["rqst"]["type"]
        rqstURL  = rqst["rqst"]["url"]
        try:
            rqstStream = rqst["rqst"]["stream"]
        except:
            rqstStream = False

        try:
            csvStream = rqst["rqst"]["csv"]
        except:
            csvStream = False
            
        myprint(1,'Request type: %s, Request URL: %s' % (rqstType, rqstURL))
        myprint(2,'Request Headers:', json.dumps(hdrs.headers, indent=2))

        errFlag = False
        
        if rqstType == 'GET':
            try:
                myprint(2,'Request Stream:', rqstStream, 'CSV Stream:', csvStream)
                r = self._session.get(rqstURL, headers=hdrs.headers, stream=rqstStream)
            except requests.exceptions.RequestException as e:
                errFlag = True
                
        elif rqstType == 'POST':
            rqstPayloadData = rqst["rqst"]["payload_data"]
                
            if rqst["rqst"]["payload_type"] == 'MULTIPART_FORM_DATA':
                myprint(2, json.dumps(dict(rqstPayloadData), indent=2))
                try:
                    r = self._session.post(rqstURL,
                                           headers=hdrs.headers,
                                           files=rqstPayloadData)
                except requests.exceptions.RequestException as e:
                    errFlag = True
            else:
                myprint(2,"payload=%s" % rqstPayloadData)
                try:
                    r = self._session.post(rqstURL,
                                           headers=hdrs.headers,
                                           data=rqstPayloadData)
                except requests.exceptions.RequestException as e:
                    errFlag = True
                
        else:	# OPTIONS
            assert(rqstType == 'OPTIONS')
            try:
                r = self._session.options(rqstURL, headers=hdrs.headers)
            except requests.exceptions.RequestException as e:
                errFlag = True

        if errFlag:
            errorMsg = 'ErRoR while retrieving information' # Dont't change the cast for ErRoR  !!!!
            myprint(0, errorMsg)
            return errorMsg

        myprint(1,'Response Code:',r.status_code)

        if r.status_code != rqst["resp"]["code"]:
            myprint(1,'Invalid Status Code: %d (expected %d). Reason: %s' % (r.status_code, rqst["resp"]["code"], r.reason))
            if rqst["returnText"]:
                return ''
            else:
                return

        myprint(2,'Response Headers:', json.dumps(dict(r.headers), indent=2))
        
        # Optional parameter "useContentDisposition" and "dumpResponse"
        fname = None	# Build output filename from response header

        try:
            useContentDisposition = rqst["resp"]["useContentDisposition"]
        except:
            myprint(2, 'Content-Disposition header not found in response')
            # Manually build output filename
            fname = datetime.datetime.now().strftime("%d%m%Y") + '.xlsx'
        else:
            if useContentDisposition:
                # Example: "Content-Disposition": "attachment;filename=20221221.xlsx",
                try:
                    cd = r.headers['Content-Disposition']
                except:
                    myprint(2, 'No Content-Disposition found in response header')
                else:
                    for item in cd.split(';'):
                        if item.startswith('filename='):
                            fname = item.split('=')[1]
                            #myprint(1, 'Using output filename:', fname)
                            break
                    
                    if not fname:
                        myprint(1, 'ERROR while parsing Content-Disposition response header')
                        # Manually build output filename
                        fname = datetime.datetime.now().strftime("%d%m%Y") + '.xlsx'

        try:
            dumpResponse = rqst["resp"]["dumpResponse"]
        except:
            myprint(2, 'No "dumpResponse" requested')
            pass
        else:
            myprint(1, 'Using output filename:', fname)
            outputFile = os.path.join(mg.moduleDirPath, fname) #rqst["resp"]["dumpResponse"])
            myprint(1, 'Using output file path:', outputFile)

            # Update the HTTP request with new output file path
            rqst['resp']['dumpResponse'] = outputFile
                    
            if rqstStream:
                if csvStream:
                    with open(outputFile, 'wb') as f:
                        for line in r.iter_lines():
                            f.write(line+'\n'.encode())
                else:
                    r.raw.decode_content = True
                    myprint(1, "Saving raw text to %s" % outputFile)
                    with open(outputFile, 'wb') as f:
                        shutil.copyfileobj(r.raw, f)
            else:
                myprint(2, "dumpToFile(%s, r.content)" % outputFile)
                dumpToFile(outputFile, r.content)
        
        # Update cookies
        if rqst["resp"]["updateCookies"]:
            self._updateCookies(r.cookies)
            
        if rqst["returnText"]:
            return r.text

        return ''
    

def parseDailyRates(filePath, dt):

    myprint(2, 'Parsing %s' % filePath)
    wb = load_workbook(filePath)
    ws = wb.active

    # rowno = 1
    # for row in ws.iter_rows():
    #    cellno = 1
    #    for cell in row:
    #        if cell.value:
    #            print(rowno,cellno,cell.value)
    #            cellno += 1
    #    rowno += 1

    myprint(1, 'Date:',ws.cell(row=10, column=11).value, '  ', 'Buying Notes Rate:', ws.cell(row=10, column=7).value)

    currency = ws.cell(row=10, column=2).value		# Currency
    code = ws.cell(row=10, column=3).value		# Code
    ratesDate = ws.cell(row=10, column=11).value	# Rates Date like "21-Dec-2022 08:54"

    # BUYING
    buyingRates = [ws.cell(row=10, column=5).value, # TT
                   ws.cell(row=10, column=6).value, # TC/DD
                   ws.cell(row=10, column=7).value  # NOTES
    ]

    # SELLING
    sellingRates = [ws.cell(row=10, column=8).value, # TT
                    ws.cell(row=10, column=9).value, # TC/DD
                    ws.cell(row=10, column=10).value # NOTES
    ]

    # Convert ratesDate into a string of form: %Y%m%d
    if ratesDate:
        d = datetime.strptime(ratesDate, '%d-%b-%Y %H:%M').strftime("%Y%m%d")
    else:
        myprint(1, 'Unable to parse ratesDate from response.')
        # If dt is null, we are parsing an existing xlsx input file
        if dt:
            myprint(1, 'Adding null entry to cache')
            d = dt
            code = 'EUR'
            buyingRates = sellingRates = ['','','']
        else:
            return

    # Add data to cache
    dataCacheDict = loadDataFromCacheFile()
    if not dataCacheDict:
        myprint(1, 'Failed to load data from local cache file. Creating new cache')
        dataCacheDict = dict()

    dataCacheDict[d] = (d, code, buyingRates, sellingRates)
    myprint(1, json.dumps(dataCacheDict, indent=4))

    # Update the data cache file
    dumpJsonToFile(mg.dataCachePath, dataCacheDict)

    if config.DEBUG:
        for key in sorted(dataCacheDict):
            myprint (1, "%s: %s" % (key, dataCacheDict[key]))


####
# Load data from local cache. If cache is older than delay, return None to force a reload
def loadDataFromCacheFile():

    if not os.path.isfile(mg.dataCachePath):	# Cache file does not exists
        return None
    
    myprint(1, 'Loading data from local cache')

    try:
        with open(mg.dataCachePath, 'r') as infile:
            data = infile.read()
            res = json.loads(data)
            return res
    except Exception as error: 
        myprint(0, f"Unable to open data cache file {mg.dataCachePath}")
        return None


def getDailyRatesFromMCBServer(forexDate):  # dd/mm/yyyy
    global cacheUpdated

    # Create session
    with requests.session() as session:
        # Create connection with MCB server
        mcb = MCBDailyRates(session)
        
        # Get information from server
        res = mcb.getDailyRates(forexDate)
        if not res:
            myprint(1, 'Cache file updated')
            cacheUpdated = True
        return res

def showDailyRatesInfo(drd):  # dd/mm/yyyy

    # Load data from local cache
    dataCacheDict = loadDataFromCacheFile()

    try:
        dailyRatesDate = datetime.strptime(drd, '%d/%m/%Y').strftime("%Y%m%d")
    except:
        myprint(0, f'Invalid/Not found input date: {drd}')

    if not dataCacheDict or not dailyRatesDate in dataCacheDict:
        myprint(1, 'Failed to load data for given date from local cache file. Retrieving data from server')
        # Read data from server
        res = getDailyRatesFromMCBServer(drd)
        if res:
            myprint(0, 'Failed to create/update local data cache')
            return -1

        dataCacheDict = loadDataFromCacheFile()
        # Assuming no error
        
        if config.DEBUG:
            t = os.path.getmtime(mg.dataCachePath)
            dt = datetime.fromtimestamp(t).strftime('%Y/%m/%d %H:%M:%S')
            myprint(1, f'Cache file updated. Last modification time: {dt}')

    for key in sorted(dataCacheDict):
        myprint (1, "%s: %s" % (key, dataCacheDict[key]))

    labels = ['Date',
              'Code',
              'Buying TT',
              'Buying TC/DD',                  
              'Buying NOTES',
              'Selling TT',
              'Selling TC/DD',                  
              'Selling NOTES'
        ]

    # example: ['20221223', 'EUR', ['45.52', '45.41', '45.09'], ['46.89', '46.89', '46.89']]

    TT = 0
    TCDD = 1
    NOTES = 2
    
    try:
        ratesDate, code, buyingList, sellingList = dataCacheDict[dailyRatesDate]
    except:
        myprint(0, f'Invalid/Not found input date: {dailyRatesDate}')
        return -1

    if config.VERBOSE:
        s = "{B}Daily Rates for : {DATE}{E} {CA}".format(
            B=color.BOLD,
            E=color.END,
            CA="(+)" if cacheUpdated else "",
            DATE=datetime.strptime(ratesDate, '%Y%m%d').strftime("%a %d %b, %Y"))
        print(s)
        
        print("Currency code: {C}".format(C=code))
        
        for i in range(2):
            s = "{L:<15}: {R:6}".format(L=labels[i+2], R=buyingList[i])
            print(s)

        # Focus on Buying Notes
        s = "{L:<15}: {B}{R:6}{E}".format(L=labels[4], R=buyingList[NOTES], B=color.BOLD, E=color.END)
        print(s)

        for i in range(3):
            s = "{L:<15}: {R:6}".format(L=labels[i+5], R=sellingList[i])
            print(s)

    else:	# Short output
        print(dataCacheDict[dailyRatesDate])

    return 0

def showHistoryRates():

    dataCacheDict = loadDataFromCacheFile()
    for key in sorted(dataCacheDict):
        print("%s: %s" % (key, dataCacheDict[key]))
